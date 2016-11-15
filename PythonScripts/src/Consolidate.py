# a script to consolidate important data from multiple .xlsx files.
import os
from openpyxl import Workbook
from openpyxl import load_workbook
import statistics

# create workbook for results
conWb = Workbook()
conWs = conWb.active
conWs.title = 'Sheet1'

headers = ['ensemble id', 'number of runs', 'median fitness', 'mode fitness', 'mean fitness', 'lowest fitness',
           'highest fitness', 'range', 'algorithms']

conWs.append(headers)

rowCount = 2

# for all excel files in current directory
for excelFileName in os.listdir(os.getcwd()):
    if not excelFileName.endswith('.xlsx'):
        continue

    print("Adding " + excelFileName + " to file.")
    # load workbook
    wb = load_workbook(filename=excelFileName)
    ws = wb.active

    # pull data from file
    enId = ws['F2'].value
    runs = ws.max_row - 1

    fitness = []
    for i in range(2, ws.max_row):
        var = ws['J' + str(i)].value
        fitness.append(var)

    fitness.sort()

    # add pulled data to final workbook
    conWs['A' + str(rowCount)] = enId
    conWs['B' + str(rowCount)] = runs
    conWs['C' + str(rowCount)] = statistics.median(fitness)

    try:
        conWs['D' + str(rowCount)] = statistics.mode(fitness)
    except statistics.StatisticsError:
        print('')
    conWs['E' + str(rowCount)] = statistics.mean(fitness)
    conWs['F' + str(rowCount)] = fitness[0]
    conWs['G' + str(rowCount)] = fitness[-1]
    conWs['H' + str(rowCount)] = fitness[-1] - fitness[0]
    conWs['I' + str(rowCount)] = ws['I2'].value

    rowCount += 1
# save workbook
conWb.save(filename='consolidatedData.xlsx')
